from __future__ import annotations

"""
SMARTPSS PLAYBACK BATCH EXPORT - FIXED DOWNLOAD FOLDER V3.3
============================================================

Automates SmartPSS Playback on Windows using calibrated screen coordinates.
Each Excel task is identified by:
    Device Name + Channel (1-13)

Workflow:
1. Focus the device search field, press Ctrl+A, and enter the device name.
2. Click the device search icon.
3. Expand the matching device.
4. Click the calibrated Channel 1-13 position.
5. Select Record / All Records / Main Stream.
6. Open Time and enter the start and end date-times as
   yyyy-mm-dd hh:mm:ss.
7. Click Playback Search exactly once, click Play, and wait 5 seconds.
8. Click the scissors icon twice with a one-second interval.
9. In Export Setup, keep the existing folder, select MP4, and confirm.
10. Confirm the prompt.
11. Monitor the fixed download directory stored in the JSON configuration.
12. Wait for a new MP4 to appear, stop growing, and be released by SmartPSS.
13. Rename the MP4, close the download window once, and continue.

Emergency stop: move the mouse pointer to the top-left corner of the primary
screen and keep it there until PyAutoGUI raises its fail-safe exception.
"""

import argparse
import ctypes
import json
import os
import re
import shutil
import sys
import time
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

CONFIG_FILE_DEFAULT = Path(__file__).with_name("smartpss_export_config.json")

REQUIRED_HEADERS = [
    "Task ID",
    "Device Name",
    "Channel",
    "Camera Name",
    "Start Time",
    "End Time",
]

MANAGED_HEADERS = [
    "Status",
    "Output File",
    "Processing Started",
    "Processing Finished",
    "Elapsed Time",
    "Notes",
]

STATUS_PROCESSING = "IN PROGRESS"
STATUS_DONE = "COMPLETED"
STATUS_ERROR = "ERROR"

VIDEO_EXTENSION = ".mp4"

pyautogui = None
pyperclip = None


def require_gui_modules() -> None:
    """Import GUI libraries only when calibration/export is actually run."""
    global pyautogui, pyperclip
    if pyautogui is not None and pyperclip is not None:
        return

    try:
        import pyautogui as _pyautogui
        import pyperclip as _pyperclip
    except ImportError as exc:
        raise SystemExit(
            "Missing GUI libraries. Install them with:\n"
            "py -m pip install pyautogui pyperclip openpyxl"
        ) from exc

    pyautogui = _pyautogui
    pyperclip = _pyperclip
    pyautogui.FAILSAFE = True
    pyautogui.PAUSE = 0.20


try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing openpyxl. Install it with:\n"
        "py -m pip install pyautogui pyperclip openpyxl"
    ) from exc


@dataclass
class ExportTask:
    excel_row: int
    task_id: Any
    device_name: str
    channel: int
    camera_name: str
    start_time: datetime
    end_time: datetime
    status: str = ""
    result_file: str = ""


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text)
    return " ".join(text.strip().split()).casefold()


def safe_filename_component(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "").strip())
    text = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", text)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text).strip("._ ")
    return text or "UNKNOWN"


def parse_excel_datetime(value: Any, field_name: str, row: int) -> datetime:
    """
    Read Excel date-time values reliably.

    Excel normally returns a datetime object even when the cell is displayed as
    "7/20/2026 7:30:00 AM". Text cells in that same US-style format are also
    accepted. Regardless of Excel display, SmartPSS receives
    yyyy-mm-dd hh:mm:ss.
    """
    if isinstance(value, datetime):
        return value.replace(microsecond=0)

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    # Normalize repeated spaces, e.g. "7/20/2026  7:30:00 AM".
    text = " ".join(str(value or "").strip().split())
    formats = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    raise ValueError(
        f"Row {row}: '{field_name}' could not be parsed. "
        "Use a real Excel date-time cell or yyyy-mm-dd hh:mm:ss."
    )


def smartpss_datetime(value: datetime) -> str:
    # SmartPSS receives the required format: yyyy-mm-dd hh:mm:ss
    return value.strftime("%Y-%m-%d %H:%M:%S")


def format_elapsed(seconds: float) -> str:
    total_seconds = max(0, int(round(seconds)))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def windows_is_admin() -> bool:
    """Return True when the current Python process is elevated on Windows."""
    if os.name != "nt":
        return False
    try:
        return bool(ctypes.windll.shell32.IsUserAnAdmin())
    except Exception:
        return False


def click(point: Iterable[int], clicks: int = 1, interval: float = 0.15) -> None:
    """
    Perform explicit left-button down/up events.

    SmartPSS uses several custom controls. Explicit mouseDown/mouseUp with
    short pauses is more reliable than a very fast pyautogui.click().
    """
    x, y = [int(v) for v in point]
    pyautogui.moveTo(x, y, duration=0.25)
    time.sleep(0.20)

    for index in range(clicks):
        pyautogui.mouseDown(button="left")
        time.sleep(0.10)
        pyautogui.mouseUp(button="left")
        if index < clicks - 1:
            time.sleep(interval)

    time.sleep(0.30)


def press_ctrl_a() -> None:
    """Send Ctrl+A with explicit key-down/key-up events and visible pauses."""
    pyautogui.keyDown("ctrl")
    time.sleep(0.15)
    pyautogui.press("a")
    time.sleep(0.15)
    pyautogui.keyUp("ctrl")
    time.sleep(0.30)


def press_ctrl_v() -> None:
    """Send Ctrl+V with explicit key-down/key-up events."""
    pyautogui.keyDown("ctrl")
    time.sleep(0.15)
    pyautogui.press("v")
    time.sleep(0.15)
    pyautogui.keyUp("ctrl")
    time.sleep(0.40)


def paste_text(
    point: Iterable[int],
    value: str,
    *,
    type_ascii: bool = False,
) -> None:
    """
    Focus a field, select all existing content, clear it, then enter new text.

    type_ascii=True is used for date-time values because they contain ASCII
    characters only and some SmartPSS date controls accept keyboard typing
    more reliably than clipboard paste.
    """
    value = str(value)
    pyperclip.copy(value)

    click(point)
    time.sleep(0.50)

    # Repeat Ctrl+A once because some SmartPSS controls consume the first
    # keyboard event while receiving focus.
    press_ctrl_a()
    press_ctrl_a()

    pyautogui.press("backspace")
    time.sleep(0.30)

    if type_ascii:
        pyautogui.write(value, interval=0.04)
        time.sleep(0.40)
    else:
        press_ctrl_v()

    time.sleep(0.40)


def paste_datetime_exact(point: Iterable[int], value: datetime) -> None:
    """
    Enter a SmartPSS date-time as one atomic clipboard paste.

    SmartPSS uses a masked date-time control. Typing the characters one by one
    can make the control auto-advance between hour/minute/second segments and
    turn 07:30:00 into 07:00:30. Pasting the complete formatted value avoids
    that segment-shifting behaviour.
    """
    formatted = smartpss_datetime(value)
    pyperclip.copy(formatted)

    click(point)
    time.sleep(0.60)
    press_ctrl_a()
    press_ctrl_a()
    pyautogui.press("backspace")
    time.sleep(0.30)
    press_ctrl_v()
    time.sleep(0.80)

    # Commit the masked field before moving to another control.
    pyautogui.press("tab")
    time.sleep(0.40)


def select_dropdown(dropdown: Iterable[int], option: Iterable[int]) -> None:
    click(dropdown)
    time.sleep(0.60)
    click(option)


def capture_point(description: str, seconds: int = 4) -> list[int]:
    print(f"\nCapture coordinate: {description}")
    input("Press Enter, then move the mouse to the target position...")
    for remaining in range(seconds, 0, -1):
        print(f"  {remaining} seconds remaining", end="\r", flush=True)
        time.sleep(1)
    point = pyautogui.position()
    print(f"  Saved: ({point.x}, {point.y})          ")
    return [point.x, point.y]


def calibrate(config_path: Path) -> None:
    print(
        "\nSMARTPSS FIXED EXPORT V3.3 - COORDINATE CALIBRATION\n"
        "========================================\n"
        "- Open SmartPSS on the Playback screen.\n"
        "- Maximize the window and keep its size unchanged.\n"
        "- Do not change resolution or Windows Display Scaling after calibration.\n"
        "- Channels must remain at fixed positions after searching and expanding a device.\n"
    )

    points: dict[str, Any] = {}

    print("\nA. DEVICE SEARCH")
    points["device_search_input"] = capture_point("device search input")
    points["device_search_button"] = capture_point("device search icon")
    points["device_expand_arrow"] = capture_point(
        "device expand arrow after search"
    )

    print(
        "\nB. 13 CHANNEL POSITIONS\n"
        "Open a device with all or nearly all channels so the channel rows are visible.\n"
        "Each position is captured independently; spacing is not inferred."
    )
    input("When the channel list is visible, press Enter...")
    channel_points: dict[str, list[int]] = {}
    for channel in range(1, 14):
        channel_points[str(channel)] = capture_point(
            f"Channel {channel} click position"
        )
    points["channels"] = channel_points

    clicks_answer = input(
        "\nHow many clicks select a channel? 1 or 2 [default 1]: "
    ).strip()
    channel_clicks = 2 if clicks_answer == "2" else 1

    print("\nC. PLAYBACK FILTERS")
    points["record_dropdown"] = capture_point("menu Record/Picture")
    input("Open Record/Picture so the Record option is visible, then press Enter...")
    points["record_option"] = capture_point("Record option")

    points["record_type_dropdown"] = capture_point("menu Record Type")
    input("Open Record Type so All Records is visible, then press Enter...")
    points["all_records_option"] = capture_point("All Records option")

    points["stream_dropdown"] = capture_point("menu Stream Type")
    input("Open Stream Type so Main Stream is visible, then press Enter...")
    points["main_stream_option"] = capture_point("Main Stream option")

    print("\nD. TIME WINDOW")
    points["time_button"] = capture_point("Time control that opens the date-time window")
    input("Click Time manually to open the small window, then press Enter...")
    points["start_datetime_input"] = capture_point("start date-time input")
    points["end_datetime_input"] = capture_point("end date-time input")

    has_apply = input(
        "Does the Time window have a separate OK/Apply button? y/N: "
    ).strip().lower() == "y"
    if has_apply:
        points["time_apply_button"] = capture_point("Time window OK/Apply button")
    else:
        points["time_apply_button"] = None

    print("\nE. SEARCH, PLAY AND SCISSORS")
    points["search_button"] = capture_point("Playback Search button")
    points["play_button"] = capture_point("Play button")
    points["scissors_button"] = capture_point("scissors icon")

    print(
        "\nF. EXPORT SETUP\n"
        "Continue manually until the Export Setup window is open."
    )
    input("When Export Setup is open, press Enter...")
    points["format_dropdown"] = capture_point("format dropdown")
    input("Open the format list so MP4 is visible, then press Enter...")
    points["mp4_option"] = capture_point("MP4 option")
    points["export_ok_button"] = capture_point("Export Setup OK button")

    print("\nG. PROMPT")
    input("Click Export Setup OK manually to display the prompt, then press Enter...")
    points["prompt_ok_button"] = capture_point("prompt OK button")

    print("\nH. CLOSE DOWNLOAD WINDOW")
    input(
        "Click the prompt OK button manually so the download/export progress window appears. "
        "When the X or Close button is visible, press Enter..."
    )
    points["download_window_close_button"] = capture_point(
        "X/Close button used after the video is complete"
    )

    fixed_download_dir_text = input(
        "\nEnter the exact fixed download directory used by SmartPSS Playback Export: "
    ).strip().strip('"')
    fixed_download_dir = Path(fixed_download_dir_text).expanduser()
    if not fixed_download_dir.exists() or not fixed_download_dir.is_dir():
        raise ValueError(
            "The fixed download directory does not exist or is not a directory: "
            f"{fixed_download_dir}"
        )

    config = {
        "fixed_download_dir": str(fixed_download_dir.resolve()),
        "points": points,
        "channel_clicks": channel_clicks,
        "deselect_channel_after_download": False,
        "delays": {
            "after_device_search_seconds": 2.0,
            "after_device_expand_seconds": 1.5,
            "after_channel_select_seconds": 1.0,
            "after_filter_select_seconds": 0.5,
            "after_time_input_seconds": 0.8,
            "after_search_seconds": 8.0,
            "after_play_seconds": 5.0,
            "between_scissors_seconds": 1.0,
            "wait_export_setup_seconds": 2.0,
            "wait_prompt_seconds": 2.0,
            "after_prompt_ok_seconds": 3.0,
            "after_download_window_close_seconds": 2.0,
            "after_return_to_search_seconds": 1.0,
            "file_poll_seconds": 10,
            "stable_checks_required": 6,
            "completion_poll_seconds": 5,
            "completion_stable_seconds": 20
        }
    }

    config_path.write_text(
        json.dumps(config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\nConfiguration file created:\n{config_path}")


def calibrate_download_close_button(config_path: Path) -> None:
    """
    Add only the download-window Close/X coordinate to an existing JSON file.
    """
    if not config_path.exists():
        raise FileNotFoundError(
            f"Configuration file not found: {config_path}\n"
            "Run --calibrate first."
        )

    config = json.loads(config_path.read_text(encoding="utf-8"))
    config.setdefault("points", {})
    config.setdefault("delays", {})

    print(
        "\nCALIBRATE DOWNLOAD-WINDOW CLOSE BUTTON ONLY\n"
        "=====================================\n"
        "Open SmartPSS and leave the download/export progress window visible.\n"
        "Move the mouse to the X or Close button used after a video completes."
    )
    config["points"]["download_window_close_button"] = capture_point(
        "download-window X/Close button"
    )
    config["delays"].setdefault(
        "after_download_window_close_seconds",
        1.5,
    )

    config_path.write_text(
        json.dumps(config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\nConfiguration updated:\n{config_path}")


def set_fixed_download_dir(config_path: Path, folder_value: str) -> None:
    """
    Store the SmartPSS default Record Path in the existing JSON config.

    This does not recalibrate any screen coordinates.
    """
    if not config_path.exists():
        raise FileNotFoundError(
            f"Configuration file not found: {config_path}\n"
            "Run --calibrate first."
        )

    folder = Path(str(folder_value).strip().strip('"')).expanduser()
    if not folder.exists() or not folder.is_dir():
        raise ValueError(
            "The fixed download directory does not exist or is not a directory: "
            f"{folder}"
        )

    config = json.loads(config_path.read_text(encoding="utf-8"))
    config["fixed_download_dir"] = str(folder.resolve())

    config_path.write_text(
        json.dumps(config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(
        "\nFixed download directory updated in JSON:\n"
        f"{config['fixed_download_dir']}"
    )


def load_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(
            f"Configuration file not found: {path}\n"
            "Open SmartPSS Playback and run --calibrate first."
        )
    config = json.loads(path.read_text(encoding="utf-8"))
    channels = config.get("points", {}).get("channels", {})
    missing = [str(i) for i in range(1, 14) if str(i) not in channels]
    if missing:
        raise ValueError("Configuration is missing channel positions: " + ", ".join(missing))
    delays = config.setdefault("delays", {})
    delays.setdefault("after_download_window_close_seconds", 2.0)
    delays.setdefault("after_return_to_search_seconds", 1.0)
    delays.setdefault("completion_poll_seconds", 5)
    delays.setdefault("completion_stable_seconds", 20)

    # Ignore unsafe legacy settings. The download window must never be
    # double-clicked because the second click can reach the Search button
    # behind the modal.
    config["double_click_download_close"] = False
    return config


def find_headers(ws: Any) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in ws[1]:
        key = normalize_text(cell.value)
        if key:
            mapping[key] = cell.column

    missing = [h for h in REQUIRED_HEADERS if normalize_text(h) not in mapping]
    if missing:
        raise ValueError("Excel is missing required columns: " + ", ".join(missing))

    next_col = ws.max_column + 1
    for header in MANAGED_HEADERS:
        key = normalize_text(header)
        if key not in mapping:
            ws.cell(row=1, column=next_col, value=header)
            mapping[key] = next_col
            next_col += 1
    return mapping


def get_value(ws: Any, row: int, columns: dict[str, int], header: str) -> Any:
    return ws.cell(row=row, column=columns[normalize_text(header)]).value


def set_value(
    ws: Any,
    row: int,
    columns: dict[str, int],
    header: str,
    value: Any,
) -> None:
    ws.cell(row=row, column=columns[normalize_text(header)], value=value)


def row_is_empty(ws: Any, row: int, columns: dict[str, int]) -> bool:
    return all(
        get_value(ws, row, columns, header) in (None, "")
        for header in REQUIRED_HEADERS
    )


def parse_channel(value: Any, row: int) -> int:
    try:
        channel = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Row {row}: Channel must be an integer from 1 to 13.") from exc
    if not 1 <= channel <= 13:
        raise ValueError(f"Row {row}: Channel must be between 1 and 13.")
    return channel


def task_from_row(ws: Any, row: int, columns: dict[str, int]) -> ExportTask:
    start = parse_excel_datetime(
        get_value(ws, row, columns, "Start Time"),
        "Start Time",
        row,
    )
    end = parse_excel_datetime(
        get_value(ws, row, columns, "End Time"),
        "End Time",
        row,
    )

    return ExportTask(
        excel_row=row,
        task_id=get_value(ws, row, columns, "Task ID"),
        device_name=str(get_value(ws, row, columns, "Device Name") or "").strip(),
        channel=parse_channel(get_value(ws, row, columns, "Channel"), row),
        camera_name=str(get_value(ws, row, columns, "Camera Name") or "").strip(),
        start_time=start,
        end_time=end,
        status=str(get_value(ws, row, columns, "Status") or "").strip(),
        result_file=str(get_value(ws, row, columns, "Output File") or "").strip(),
    )


def validate_task(task: ExportTask) -> None:
    if not task.device_name:
        raise ValueError(f"Row {task.excel_row}: Device Name is required.")
    if not task.camera_name:
        raise ValueError(f"Row {task.excel_row}: Camera Name is required.")
    if task.end_time <= task.start_time:
        raise ValueError(
            f"Row {task.excel_row}: End Time must be later than Start Time."
        )


def output_locations(
    task: ExportTask,
    config: dict[str, Any],
) -> tuple[Path, Path]:
    """
    Return the single fixed SmartPSS download folder and the final file name.

    SmartPSS downloads every export into fixed_download_dir. Python then
    renames the completed MP4 inside that same folder.
    """
    folder_value = str(config.get("fixed_download_dir") or "").strip()
    if not folder_value:
        raise RuntimeError(
            "The JSON configuration has no fixed_download_dir. Run "
            "--set-download-dir with the exact SmartPSS Playback Export folder."
        )

    export_dir = Path(folder_value).expanduser()
    if not export_dir.exists() or not export_dir.is_dir():
        raise FileNotFoundError(
            f"The fixed download directory does not exist: {export_dir}"
        )

    target = export_dir / (
        f"{safe_filename_component(task.task_id)}_"
        f"{safe_filename_component(task.device_name)}_"
        f"CH{task.channel:02d}_"
        f"{safe_filename_component(task.camera_name)}_"
        f"{task.start_time:%Y%m%d_%H%M%S}-"
        f"{task.end_time:%Y%m%d_%H%M%S}.mp4"
    )
    return export_dir.resolve(), target.resolve()


def mp4_snapshot(folder: Path) -> dict[Path, tuple[int, int]]:
    """
    Return all MP4 files recursively, case-insensitively.

    Path.rglob("*.mp4") can miss uppercase extensions on some filesystems.
    Scanning all files and testing suffix.lower() avoids that problem.
    Metadata contains file size and nanosecond modification time.
    """
    snapshot: dict[Path, tuple[int, int]] = {}
    if not folder.exists():
        return snapshot

    for path in folder.rglob("*"):
        try:
            if not path.is_file() or path.suffix.lower() != VIDEO_EXTENSION:
                continue
            stat = path.stat()
            snapshot[path.resolve()] = (stat.st_size, stat.st_mtime_ns)
        except OSError:
            continue
    return snapshot


def file_is_exclusively_available(path: Path) -> bool:
    """
    Check whether SmartPSS has released the completed file.

    On Windows, request an exclusive read handle (share mode 0). If SmartPSS
    still holds the file open, CreateFileW normally fails. On non-Windows
    systems, fall back to a simple read-open check.
    """
    if os.name != "nt":
        try:
            with path.open("rb") as file:
                file.read(1)
            return True
        except OSError:
            return False

    GENERIC_READ = 0x80000000
    OPEN_EXISTING = 3
    FILE_ATTRIBUTE_NORMAL = 0x00000080
    INVALID_HANDLE_VALUE = ctypes.c_void_p(-1).value

    handle = ctypes.windll.kernel32.CreateFileW(
        str(path),
        GENERIC_READ,
        0,  # no sharing: fail while another process still owns the file
        None,
        OPEN_EXISTING,
        FILE_ATTRIBUTE_NORMAL,
        None,
    )
    if handle == INVALID_HANDLE_VALUE:
        return False

    ctypes.windll.kernel32.CloseHandle(handle)
    return True


def wait_for_completed_mp4(
    folder: Path,
    before: dict[Path, tuple[int, int]],
    timeout_minutes: int,
    poll_seconds: int,
    stable_seconds_required: int,
    task_started_perf: float,
    export_started_wall: float,
) -> Path:
    """
    Wait for a new/changed MP4, stable size, and released file handle.

    Completion requires:
    - MP4 is new or changed relative to the pre-export snapshot.
    - File was modified after this export started.
    - Size and mtime remain unchanged for stable_seconds_required.
    - SmartPSS no longer locks the file.
    """
    deadline = time.time() + timeout_minutes * 60
    candidate: Path | None = None
    last_signature: tuple[int, int] | None = None
    stable_seconds = 0
    first_seen_wall: float | None = None

    print(
        "\n  Monitoring fixed download directory:\n"
        f"  {folder}\n"
        "  Waiting for a new MP4..."
    )

    while time.time() < deadline:
        current = mp4_snapshot(folder)
        changed: list[Path] = []

        for path, metadata in current.items():
            size, mtime_ns = metadata
            mtime_seconds = mtime_ns / 1_000_000_000

            is_new_or_changed = (
                path not in before or before.get(path) != metadata
            )
            is_fresh_for_this_export = mtime_seconds >= export_started_wall - 5

            if is_new_or_changed and is_fresh_for_this_export and size > 0:
                changed.append(path)

        elapsed_text = format_elapsed(time.perf_counter() - task_started_perf)

        if not changed:
            print(
                f"  No new MP4 yet | elapsed {elapsed_text}",
                end="\r",
                flush=True,
            )
            time.sleep(poll_seconds)
            continue

        newest = max(
            changed,
            key=lambda path: (
                current[path][1],
                current[path][0],
            ),
        )
        signature = current[newest]
        size = signature[0]

        if newest != candidate:
            candidate = newest
            last_signature = signature
            stable_seconds = 0
            first_seen_wall = time.time()
        elif signature == last_signature:
            stable_seconds += poll_seconds
        else:
            last_signature = signature
            stable_seconds = 0

        released = False
        if stable_seconds >= stable_seconds_required:
            released = file_is_exclusively_available(newest)

        seen_for = (
            int(time.time() - first_seen_wall)
            if first_seen_wall is not None
            else 0
        )
        lock_text = (
            "file released"
            if released
            else "writing/file locked"
        )

        print(
            f"  Found: {newest.name} | "
            f"{size / (1024 * 1024):.1f} MB | "
            f"stable {stable_seconds}/{stable_seconds_required}s | "
            f"{lock_text} | "
            f"visible for {seen_for}s | "
            f"elapsed {elapsed_text}",
            end="\r",
            flush=True,
        )

        if stable_seconds >= stable_seconds_required and released:
            print(
                "\n  MP4 COMPLETION CONFIRMED: "
                f"{newest.name}"
            )
            return newest

        time.sleep(poll_seconds)

    raise TimeoutError(
        f"Could not confirm a completed MP4 within {timeout_minutes} minutes in: "
        f"{folder}"
    )


def move_to_target(downloaded: Path, target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    if downloaded.resolve() == target.resolve():
        return target

    if target.exists():
        backup = target.with_name(
            f"{target.stem}_OLD_{datetime.now():%Y%m%d_%H%M%S}{target.suffix}"
        )
        target.rename(backup)

    try:
        downloaded.replace(target)
    except OSError:
        shutil.move(str(downloaded), str(target))
    return target


def select_device_and_channel(task: ExportTask, config: dict[str, Any]) -> None:
    points = config["points"]
    delays = config["delays"]

    # 1. Find the device.
    paste_text(points["device_search_input"], task.device_name)
    click(points["device_search_button"])
    time.sleep(float(delays["after_device_search_seconds"]))

    # 2. Expand the device.
    click(points["device_expand_arrow"])
    time.sleep(float(delays["after_device_expand_seconds"]))

    # 3. Select the calibrated Channel 1-13 position.
    channel_point = points["channels"][str(task.channel)]
    click(
        channel_point,
        clicks=int(config.get("channel_clicks", 1)),
    )
    time.sleep(float(delays["after_channel_select_seconds"]))


def set_playback_filters(config: dict[str, Any]) -> None:
    points = config["points"]
    delay = float(config["delays"]["after_filter_select_seconds"])

    select_dropdown(points["record_dropdown"], points["record_option"])
    time.sleep(delay)
    select_dropdown(
        points["record_type_dropdown"],
        points["all_records_option"],
    )
    time.sleep(delay)
    select_dropdown(points["stream_dropdown"], points["main_stream_option"])
    time.sleep(delay)


def set_time_range(task: ExportTask, config: dict[str, Any]) -> None:
    points = config["points"]

    click(points["time_button"])
    time.sleep(0.8)

    # Paste the complete value atomically. Do not type the characters one by
    # one because the SmartPSS masked control can shift minute/second segments.
    paste_datetime_exact(points["start_datetime_input"], task.start_time)
    paste_datetime_exact(points["end_datetime_input"], task.end_time)

    apply_button = points.get("time_apply_button")
    if apply_button:
        click(apply_button)

    time.sleep(float(config["delays"]["after_time_input_seconds"]))


def run_export_ui(
    task: ExportTask,
    config: dict[str, Any],
) -> None:
    points = config["points"]
    delays = config["delays"]

    select_device_and_channel(task, config)
    set_playback_filters(config)
    set_time_range(task, config)

    print("  Clicking Playback Search exactly once...")
    click(points["search_button"], clicks=1)
    print("  One Search click sent; no second click.")
    time.sleep(float(delays["after_search_seconds"]))

    click(points["play_button"])
    time.sleep(float(delays["after_play_seconds"]))

    click(points["scissors_button"])
    time.sleep(float(delays["between_scissors_seconds"]))
    click(points["scissors_button"])
    time.sleep(float(delays["wait_export_setup_seconds"]))

    # Export Setup: SmartPSS already contains the fixed Playback Export folder.
    # Do not click or modify the Path field.
    select_dropdown(points["format_dropdown"], points["mp4_option"])
    click(points["export_ok_button"])
    time.sleep(float(delays["wait_prompt_seconds"]))

    # Prompt.
    click(points["prompt_ok_button"])
    time.sleep(float(delays["after_prompt_ok_seconds"]))


def close_download_window(config: dict[str, Any]) -> None:
    """Close the SmartPSS download/export-progress window."""
    point = config.get("points", {}).get("download_window_close_button")
    if not point:
        raise RuntimeError(
            "The configuration has no download-window close-button coordinate. "
            "Run: py smartpss_playback_export.py "
            "--calibrate-close-button"
        )

    print(
        "  MP4 is complete; clicking ONCE to close the "
        "SmartPSS download window..."
    )

    # Critical: exactly one physical click. A second click after the modal
    # closes can land on the Playback Search button underneath and cancel
    # the newly loaded search result.
    click(point, clicks=1)

    time.sleep(
        float(
            config.get("delays", {}).get(
                "after_download_window_close_seconds",
                2.0,
            )
        )
    )
    print("  One close click sent; waiting for the window to close.")


def return_to_device_search(config: dict[str, Any]) -> None:
    """
    Return the mouse and keyboard focus to the device-search input.

    This replaces the old post-download channel click. The next Excel row
    therefore always starts from the same known UI position.
    """
    point = config.get("points", {}).get("device_search_input")
    if not point:
        raise RuntimeError(
            "The configuration has no device-search input coordinate."
        )

    print("  Moving the mouse back to the device search field...")
    pyautogui.moveTo(
        int(point[0]),
        int(point[1]),
        duration=0.35,
    )

    # Do not click here. The next task will focus this field exactly once
    # inside paste_text(). This prevents any extra click from affecting a
    # control underneath a recently closed modal window.
    time.sleep(
        float(
            config.get("delays", {}).get(
                "after_return_to_search_seconds",
                1.0,
            )
        )
    )
    print(
        "  Mouse returned to the device search field; "
        "no click was sent."
    )


def process_task(
    task: ExportTask,
    config: dict[str, Any],
    timeout_minutes: int,
    task_started_perf: float,
) -> tuple[Path, float]:
    validate_task(task)
    export_dir, target_file = output_locations(task, config)
    before = mp4_snapshot(export_dir)

    print(
        f"\nTask ID {task.task_id} | {task.device_name} | "
        f"Channel {task.channel} | {task.camera_name}\n"
        f"  {smartpss_datetime(task.start_time)} -> "
        f"{smartpss_datetime(task.end_time)}\n"
        f"  Fixed download folder: {export_dir}"
    )

    run_export_ui(task, config)
    export_started_wall = time.time()

    downloaded = wait_for_completed_mp4(
        folder=export_dir,
        before=before,
        timeout_minutes=timeout_minutes,
        poll_seconds=int(
            config["delays"].get("completion_poll_seconds", 5)
        ),
        stable_seconds_required=int(
            config["delays"].get("completion_stable_seconds", 20)
        ),
        task_started_perf=task_started_perf,
        export_started_wall=export_started_wall,
    )
    final_file = move_to_target(downloaded, target_file)

    # Start the next Excel row only after the completed download window is
    # closed.
    close_download_window(config)

    # Do not click the selected channel again. After closing the download
    # window, return to the exact starting point of the next loop.
    return_to_device_search(config)

    elapsed_seconds = time.perf_counter() - task_started_perf
    print(f"  Video completed in {format_elapsed(elapsed_seconds)}")
    return final_file, elapsed_seconds


def make_result_workbook(source: Path, requested_result: str | None) -> Path:
    if requested_result:
        result = Path(requested_result).expanduser()
    else:
        result = source.with_name(f"{source.stem}_result{source.suffix}")

    if result.resolve() == source.resolve():
        backup = source.with_name(
            f"{source.stem}_backup_{datetime.now():%Y%m%d_%H%M%S}{source.suffix}"
        )
        shutil.copy2(source, backup)
        return source

    if not result.exists():
        shutil.copy2(source, result)
    return result


def completed_and_file_exists(task: ExportTask) -> bool:
    return (
        normalize_text(task.status) == normalize_text(STATUS_DONE)
        and bool(task.result_file)
        and Path(task.result_file).exists()
    )


def save_error_screenshot(result_excel: Path, task: ExportTask) -> str:
    try:
        folder = result_excel.parent / "smartpss_error_screenshots"
        folder.mkdir(parents=True, exist_ok=True)
        filename = (
            f"row_{task.excel_row}_stt_{safe_filename_component(task.task_id)}_"
            f"{datetime.now():%Y%m%d_%H%M%S}.png"
        )
        path = folder / filename
        pyautogui.screenshot(str(path))
        return str(path)
    except Exception:
        return ""


def validate_excel(excel_path: Path) -> tuple[int, list[str]]:
    workbook = load_workbook(excel_path)
    ws = workbook["Task_List"] if "Task_List" in workbook.sheetnames else workbook.active
    columns = find_headers(ws)

    count = 0
    errors: list[str] = []
    for row in range(2, ws.max_row + 1):
        if row_is_empty(ws, row, columns):
            continue
        try:
            task = task_from_row(ws, row, columns)
            validate_task(task)
            count += 1
        except Exception as exc:
            errors.append(str(exc))
    return count, errors


def run_batch(args: argparse.Namespace) -> None:
    source = Path(args.excel).expanduser()
    if not source.exists():
        raise FileNotFoundError(f"Excel file not found: {source}")

    config = load_config(Path(args.config).expanduser())
    fixed_dir_text = str(config.get("fixed_download_dir") or "").strip()
    if not fixed_dir_text:
        raise RuntimeError(
            "The JSON configuration has no fixed_download_dir. Run: "
            "py smartpss_playback_export.py "
            "--set-download-dir \"D:\\SmartPSS_Exports\""
        )
    fixed_dir = Path(fixed_dir_text).expanduser()
    if not fixed_dir.exists() or not fixed_dir.is_dir():
        raise FileNotFoundError(
            f"The fixed download directory in JSON does not exist: {fixed_dir}"
        )

    if not config.get("points", {}).get("download_window_close_button"):
        raise RuntimeError(
            "The JSON configuration has no download-window close-button coordinate. "
            "Open the SmartPSS download window, then run: "
            "py smartpss_playback_export.py --calibrate-close-button"
        )
    result = make_result_workbook(source, args.result_excel)

    workbook = load_workbook(result)
    ws = workbook["Task_List"] if "Task_List" in workbook.sheetnames else workbook.active
    columns = find_headers(ws)
    workbook.save(result)

    privilege_text = (
        "Python is running with Administrator privileges."
        if windows_is_admin()
        else (
            "WARNING: Python is not running with Administrator privileges. "
            "If SmartPSS was opened with Run as administrator, close Terminal "
            "and reopen Terminal with Run as administrator; otherwise Windows "
            "may allow pointer movement while blocking clicks and keyboard input."
        )
    )

    print(
        f"\nResult/status workbook:\n{result}\n\n"
        f"{privilege_text}\n"
        f"Fixed SmartPSS download directory being monitored:\n{fixed_dir}\n"
        "Close the Excel workbook before running.\n"
        "Do not use the mouse or keyboard while automation is running.\n"
        "Move the pointer to the top-left corner for an emergency stop."
    )
    input("\nPress Enter to start after a 10-second delay...")
    time.sleep(10)

    processed = 0

    for row in range(2, ws.max_row + 1):
        if row_is_empty(ws, row, columns):
            continue
        if args.max_tasks is not None and processed >= args.max_tasks:
            break

        task: ExportTask | None = None
        task_started_perf: float | None = None
        try:
            task = task_from_row(ws, row, columns)

            if args.only_task_id is not None and str(task.task_id) != str(args.only_task_id):
                continue

            if completed_and_file_exists(task) and not args.force:
                print(f"Skipping Task ID {task.task_id}: already completed.")
                continue

            if (
                normalize_text(task.status) == normalize_text(STATUS_ERROR)
                and not args.retry_errors
                and not args.force
            ):
                print(
                    f"Skipping Task ID {task.task_id}: status is ERROR. "
                    "Use --retry-errors to run it again."
                )
                continue

            task_started_perf = time.perf_counter()
            validate_task(task)
            set_value(ws, row, columns, "Status", STATUS_PROCESSING)
            set_value(ws, row, columns, "Processing Started", datetime.now())
            set_value(ws, row, columns, "Processing Finished", None)
            set_value(ws, row, columns, "Elapsed Time", None)
            set_value(ws, row, columns, "Notes", None)
            workbook.save(result)

            output, elapsed_seconds = process_task(
                task,
                config,
                args.timeout_minutes,
                task_started_perf,
            )

            set_value(ws, row, columns, "Status", STATUS_DONE)
            set_value(ws, row, columns, "Output File", str(output))
            set_value(ws, row, columns, "Processing Finished", datetime.now())
            set_value(
                ws,
                row,
                columns,
                "Elapsed Time",
                format_elapsed(elapsed_seconds),
            )
            set_value(
                ws,
                row,
                columns,
                "Notes",
                "The MP4 appeared, stopped growing, and was released. "
                f"Total elapsed time: {format_elapsed(elapsed_seconds)}.",
            )
            workbook.save(result)
            processed += 1

        except pyautogui.FailSafeException:
            if task is not None:
                set_value(ws, row, columns, "Status", STATUS_ERROR)
                set_value(ws, row, columns, "Processing Finished", datetime.now())
                if task_started_perf is not None:
                    set_value(
                        ws,
                        row,
                        columns,
                        "Elapsed Time",
                        format_elapsed(time.perf_counter() - task_started_perf),
                    )
                set_value(ws, row, columns, "Notes", "Stopped by the PyAutoGUI fail-safe.")
                workbook.save(result)
            raise

        except Exception as exc:
            message = f"{type(exc).__name__}: {exc}"
            print(f"\nERROR on Excel row {row}: {message}", file=sys.stderr)

            if task is None:
                try:
                    task = ExportTask(row, row, "", 1, "", datetime.now(), datetime.now())
                except Exception:
                    task = None

            screenshot = save_error_screenshot(result, task) if task else ""
            note = message
            if screenshot:
                note += f" | Screenshot: {screenshot}"

            set_value(ws, row, columns, "Status", STATUS_ERROR)
            set_value(ws, row, columns, "Processing Finished", datetime.now())
            if task_started_perf is not None:
                elapsed_on_error = time.perf_counter() - task_started_perf
                set_value(
                    ws,
                    row,
                    columns,
                    "Elapsed Time",
                    format_elapsed(elapsed_on_error),
                )
                note += f" | Elapsed before error: {format_elapsed(elapsed_on_error)}"
            set_value(ws, row, columns, "Notes", note)
            workbook.save(result)
            processed += 1

            if not args.continue_on_error:
                raise
            time.sleep(5)

    print(f"\nFinished. Review the workbook:\n{result}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Batch-export SmartPSS Playback MP4 files to one fixed directory using Excel tasks and Channel 1-13."
    )
    parser.add_argument(
        "--calibrate",
        action="store_true",
        help="Calibrate all SmartPSS coordinates, including 13 channel positions.",
    )
    parser.add_argument(
        "--calibrate-close-button",
        action="store_true",
        help=(
            "Add or update only the download-window close-button coordinate in the existing JSON; "
            "do not recalibrate the 13 channels."
        ),
    )
    parser.add_argument(
        "--set-download-dir",
        metavar="FOLDER",
        help=(
            "Store the fixed SmartPSS Playback Export folder in the existing JSON "
            "without recalibrating coordinates."
        ),
    )
    parser.add_argument(
        "--config",
        default=str(CONFIG_FILE_DEFAULT),
        help="JSON coordinate configuration file.",
    )
    parser.add_argument("--excel", help="Excel task workbook.")
    parser.add_argument(
        "--result-excel",
        help="Result/status workbook; default is <name>_result.xlsx.",
    )
    parser.add_argument(
        "--validate",
        action="store_true",
        help="Validate Excel data only; do not control SmartPSS.",
    )
    parser.add_argument(
        "--timeout-minutes",
        type=int,
        default=90,
        help="Maximum wait time per video; default 90 minutes.",
    )
    parser.add_argument(
        "--max-tasks",
        type=int,
        help="Process at most N tasks; use 1 for a test run.",
    )
    parser.add_argument("--only-task-id", help="Process only the specified Task ID.")
    parser.add_argument(
        "--retry-errors",
        action="store_true",
        help="Retry rows whose Status is ERROR.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Run rows again even when Status is COMPLETED.",
    )
    parser.add_argument(
        "--continue-on-error",
        action="store_true",
        help="Record the error and continue with the next row.",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    try:
        if args.calibrate:
            require_gui_modules()
            calibrate(Path(args.config).expanduser())
            return 0

        if args.calibrate_close_button:
            require_gui_modules()
            calibrate_download_close_button(
                Path(args.config).expanduser()
            )
            return 0

        if args.set_download_dir:
            set_fixed_download_dir(
                Path(args.config).expanduser(),
                args.set_download_dir,
            )
            return 0

        if not args.excel:
            parser.error(
                "Provide --excel, --calibrate, "
                "--calibrate-close-button, or --set-download-dir."
            )

        excel_path = Path(args.excel).expanduser()
        if args.validate:
            valid_count, errors = validate_excel(excel_path)
            print(f"Valid task rows: {valid_count}")
            if errors:
                print("\nValidation errors:")
                for error in errors:
                    print(f"- {error}")
                return 1
            print("Excel workbook is valid.")
            return 0

        require_gui_modules()
        run_batch(args)
        return 0

    except KeyboardInterrupt:
        print("\nStopped from the keyboard.")
        return 130
    except Exception as exc:
        print(f"\nError: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
